#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================
  AGENT DE TRADING CRYPTO — PAPER TRADING (ARGENT FICTIF)
=============================================================
Ce script surveille les nouvelles crypto toutes les 15 minutes,
analyse leur sentiment avec l'IA Claude (Anthropic), et simule
des décisions de trading BTC avec un capital fictif de 300 USDT.

⚠️  IMPORTANT : Aucun ordre réel n'est passé sur Binance.
               C'est uniquement de la simulation (paper trading).
"""

# ─────────────────────────────────────────────────────────────
#  IMPORTATION DES BIBLIOTHÈQUES
# ─────────────────────────────────────────────────────────────
import os           # Lire les variables d'environnement
import sys          # Lire les arguments passés au script (ex: --once)
import time         # Faire des pauses entre les cycles
import json         # Sauvegarder des données dans des fichiers texte
import feedparser   # Lire les flux RSS (CoinDesk, Cointelegraph)
import requests     # Récupérer le prix BTC (CoinGecko / Binance)
import anthropic    # Utiliser l'IA Claude pour analyser les news
import openpyxl     # Lire et écrire des fichiers Excel
from datetime import datetime, timezone   # Gérer les dates et heures
from email.utils import parsedate_to_datetime  # Parser les dates RSS
from dotenv import load_dotenv  # Lire le fichier .env (clés secrètes)

# ─────────────────────────────────────────────────────────────
#  CHARGEMENT DES CLÉS API DEPUIS LE FICHIER .env
# ─────────────────────────────────────────────────────────────
load_dotenv()

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
BINANCE_API_KEY   = os.getenv("BINANCE_API_KEY")
BINANCE_SECRET    = os.getenv("BINANCE_SECRET")

# ─────────────────────────────────────────────────────────────
#  PARAMÈTRES GÉNÉRAUX
# ─────────────────────────────────────────────────────────────
CAPITAL_DEPART_USDT      = 300.0
INTERVALLE_SECONDES      = 15 * 60
FICHIER_EXCEL            = "trading_journal.xlsx"
FICHIER_CACHE_NEWS       = "cache_news.json"
FICHIER_PORTFOLIO        = "portfolio_state.json"
FICHIER_HISTORIQUE_TRADES = "trades_history.json"
CACHE_EXPIRY_JOURS       = 7

COOLDOWN_POST_TRADE_S    = 4 * 3600   # 4h minimum entre trade opposé
MAX_TRADES_PAR_JOUR      = 4          # Maximum de trades par 24h
FILTRE_AGE_MAX_SECONDES  = 6 * 3600   # Ignorer les news de plus de 6h

# Sources RSS
RSS_FEEDS = [
    "https://www.coindesk.com/arc/outboundfeeds/rss/",
    "https://cointelegraph.com/rss",
]

MOTS_SPAM = {
    "casino", "presale", "pre-sale", "100x", "alphapepe", "pepeto",
    "apemars", "gambling", "airdrop", "giveaway", "prize", "sponsor",
    "advertisement", "promoted", "pepe", "shib", "doge", "meme coin",
    "gem", "moonshot",
}

MODELE_HAIKU  = "claude-haiku-4-5-20251001"
MODELE_SONNET = "claude-sonnet-4-6"

SEUIL_NEWS_PERTINENTES = 2

FRAIS_TRADING         = 0.001
STOP_LOSS_PCT         = 0.07
TAKE_PROFIT_PCT       = 0.12
MIN_PNL_ATTENDU_USDT  = 1.50   # P&L net minimum pour ouvrir/fermer une position (5× frais 0,30 USDT)

LIMITE_COUT_QUOTIDIEN_USD = 1.00
FICHIER_TOKEN_USAGE       = "token_usage.json"
COUTS_PAR_TOKEN = {
    "claude-haiku-4-5-20251001": {"input": 0.80e-6, "output": 4.00e-6},
    "claude-sonnet-4-6":         {"input": 3.00e-6, "output": 15.00e-6},
}


# ═════════════════════════════════════════════════════════════
#  FONCTIONS UTILITAIRES
# ═════════════════════════════════════════════════════════════

class LimiteQuotidienneAtteinte(Exception):
    """Levée quand le coût API Claude dépasse LIMITE_COUT_QUOTIDIEN_USD sur la journée UTC."""


def log(message: str):
    """Affiche un message dans la console avec la date et l'heure."""
    heure = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{heure}] {message}")


# ─────────────────────────────────────────────────────────────
#  GESTION DU PORTEFEUILLE FICTIF
# ─────────────────────────────────────────────────────────────

def charger_portfolio() -> dict:
    """
    Charge l'état du portefeuille depuis le fichier JSON.
    Si le fichier n'existe pas (premier lancement), crée un portefeuille vide.
    """
    if os.path.exists(FICHIER_PORTFOLIO):
        with open(FICHIER_PORTFOLIO, "r") as f:
            data = json.load(f)
        data.setdefault("prix_initial_bh",       0.0)
        data.setdefault("dernier_trade_timestamp", 0.0)
        data.setdefault("dernier_trade_type",      "")
        data.setdefault("trades_aujourd_hui",      0)
        data.setdefault("date_trades",             "")
        return data

    return {
        "usdt_disponible":         CAPITAL_DEPART_USDT,
        "btc_en_stock":            0.0,
        "en_position":             False,
        "prix_achat_btc":          0.0,
        "prix_initial_bh":         0.0,
        "dernier_trade_timestamp": 0.0,
        "dernier_trade_type":      "",
        "trades_aujourd_hui":      0,
        "date_trades":             "",
    }


def sauvegarder_portfolio(portefeuille: dict):
    """Sauvegarde l'état du portefeuille dans le fichier JSON."""
    with open(FICHIER_PORTFOLIO, "w") as f:
        json.dump(portefeuille, f, indent=2)


# ─────────────────────────────────────────────────────────────
#  GESTION DE L'HISTORIQUE DES TRADES
# ─────────────────────────────────────────────────────────────

def charger_historique_trades() -> list:
    """Charge l'historique des trades depuis trades_history.json."""
    if not os.path.exists(FICHIER_HISTORIQUE_TRADES):
        return []
    with open(FICHIER_HISTORIQUE_TRADES, "r") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return []


def sauvegarder_historique_trades(historique: list):
    """Sauvegarde l'historique des trades dans trades_history.json."""
    with open(FICHIER_HISTORIQUE_TRADES, "w") as f:
        json.dump(historique, f, indent=2, default=str)


# ─────────────────────────────────────────────────────────────
#  GESTION DU CACHE DES NEWS
# ─────────────────────────────────────────────────────────────

def charger_cache_news() -> dict:
    """
    Charge le cache des news : dict {url: timestamp_unix}.
    Expire automatiquement les entrées de plus de CACHE_EXPIRY_JOURS jours.
    """
    if not os.path.exists(FICHIER_CACHE_NEWS):
        return {}
    with open(FICHIER_CACHE_NEWS, "r") as f:
        try:
            cache = json.load(f)
        except json.JSONDecodeError:
            return {}
    if isinstance(cache, list):
        return {}
    limite = time.time() - CACHE_EXPIRY_JOURS * 86400
    return {url: ts for url, ts in cache.items() if ts > limite}


def sauvegarder_cache_news(cache: dict):
    """Sauvegarde le cache {url: timestamp_unix} dans le fichier JSON."""
    with open(FICHIER_CACHE_NEWS, "w") as f:
        json.dump(cache, f)


# ─────────────────────────────────────────────────────────────
#  ÉTAPE 1 : RÉCUPÉRATION DES NEWS (RSS CoinDesk + Cointelegraph)
# ─────────────────────────────────────────────────────────────

def _contient_spam(texte: str) -> bool:
    """Retourne True si le texte contient un mot de la liste MOTS_SPAM."""
    texte_lower = texte.lower()
    return any(mot in texte_lower for mot in MOTS_SPAM)


def _age_article_secondes(pub_date_str: str) -> float:
    """Retourne l'âge en secondes d'un article depuis sa pubDate RSS."""
    if not pub_date_str:
        return 0.0
    try:
        dt = parsedate_to_datetime(pub_date_str)
        return (datetime.now(timezone.utc) - dt).total_seconds()
    except Exception:
        return 0.0


def recuperer_nouvelles_news() -> list:
    """
    Lit les flux RSS CoinDesk et Cointelegraph, filtre le spam,
    ignore les articles de plus de 6h, déduplique par titre normalisé
    et compare avec le cache pour ne retourner que les articles nouveaux.
    """
    log("Vérification des news RSS (CoinDesk + Cointelegraph)...")

    tous_articles = []
    for feed_url in RSS_FEEDS:
        try:
            feed = feedparser.parse(feed_url)
            for entry in feed.entries:
                titre = entry.get("title", "").strip()
                lien  = entry.get("link", "").strip()
                if not titre or not lien:
                    continue
                tous_articles.append({
                    "title":     titre,
                    "link":      lien,
                    "source_id": feed.feed.get("title", feed_url),
                    "pubDate":   entry.get("published", ""),
                    "summary":   entry.get("summary", ""),
                })
        except Exception as e:
            log(f"⚠️  Erreur lecture RSS {feed_url} : {e}")

    # Filtre âge : ignorer les articles de plus de 6h
    articles_recents = [
        a for a in tous_articles
        if _age_article_secondes(a.get("pubDate", "")) <= FILTRE_AGE_MAX_SECONDES
    ]
    nb_filtres_age = len(tous_articles) - len(articles_recents)
    if nb_filtres_age > 0:
        log(f"🕐 {nb_filtres_age} article(s) ignoré(s) car publiés il y a plus de 6h")

    # Filtre anti-spam sur titre et résumé
    articles_filtres = [
        a for a in articles_recents
        if not _contient_spam(a["title"]) and not _contient_spam(a.get("summary", ""))
    ]

    # Déduplique par titre normalisé
    titres_vus: set = set()
    articles_uniques = []
    for article in articles_filtres:
        titre_norm = article["title"].lower().strip()
        if titre_norm not in titres_vus:
            titres_vus.add(titre_norm)
            articles_uniques.append(article)

    # Filtre anti-doublons via le cache
    cache = charger_cache_news()
    maintenant = time.time()

    nouvelles_news = []
    for article in articles_uniques:
        url = article["link"]
        if url not in cache:
            nouvelles_news.append(article)
            cache[url] = maintenant

    sauvegarder_cache_news(cache)

    if nouvelles_news:
        log(f"✅ {len(nouvelles_news)} nouvelle(s) news détectée(s) "
            f"({len(articles_filtres)} après filtre spam, {len(articles_uniques)} après dédup)")
    else:
        log("Aucune nouvelle news depuis la dernière vérification. En attente...")

    return nouvelles_news


# ─────────────────────────────────────────────────────────────
#  ÉTAPE 2 : RÉCUPÉRATION DU PRIX BTC + DONNÉES MARCHÉ
# ─────────────────────────────────────────────────────────────

def recuperer_prix_btc() -> float:
    """
    Récupère le prix actuel du Bitcoin.
    Source primaire  : CoinGecko
    Source de backup : Binance ticker public
    """
    try:
        url = "https://api.coingecko.com/api/v3/simple/price?ids=bitcoin&vs_currencies=usd"
        reponse = requests.get(url, timeout=10)
        reponse.raise_for_status()
        prix = float(reponse.json()["bitcoin"]["usd"])
        log(f"💰 Prix actuel du BTC : {prix:,.2f} USD (CoinGecko)")
        return prix
    except Exception as e:
        log(f"⚠️  CoinGecko indisponible ({e}) — tentative Binance...")

    try:
        url = "https://api.binance.com/api/v3/ticker/price?symbol=BTCUSDT"
        reponse = requests.get(url, timeout=10)
        reponse.raise_for_status()
        prix = float(reponse.json()["price"])
        log(f"💰 Prix actuel du BTC : {prix:,.2f} USD (Binance backup)")
        return prix
    except Exception as e:
        log(f"⚠️  Binance indisponible ({e}) — prix BTC introuvable.")
        return 0.0


def recuperer_donnees_marche_btc() -> dict:
    """
    Récupère les données de marché BTC depuis CoinGecko :
    - Variation 1h, 24h, 7j (en %)
    - Prix min et max des dernières 24h

    Retourne un dict avec les clés : prix, var_1h, var_24h, var_7j, min_24h, max_24h.
    Retourne un dict vide en cas d'erreur.
    """
    try:
        url = (
            "https://api.coingecko.com/api/v3/coins/markets"
            "?vs_currency=usd&ids=bitcoin"
            "&price_change_percentage=1h,24h,7d"
        )
        reponse = requests.get(url, timeout=10)
        reponse.raise_for_status()
        data = reponse.json()[0]
        return {
            "prix":    data.get("current_price", 0.0),
            "var_1h":  data.get("price_change_percentage_1h_in_currency", 0.0) or 0.0,
            "var_24h": data.get("price_change_percentage_24h_in_currency", 0.0) or 0.0,
            "var_7j":  data.get("price_change_percentage_7d_in_currency", 0.0) or 0.0,
            "min_24h": data.get("low_24h", 0.0) or 0.0,
            "max_24h": data.get("high_24h", 0.0) or 0.0,
        }
    except Exception as e:
        log(f"⚠️  Données marché CoinGecko indisponibles ({e})")
        return {}


def recuperer_fear_greed() -> dict:
    """
    Récupère le Fear & Greed Index crypto via alternative.me (API gratuite).
    Retourne {"valeur": int, "classification": str} ou {} en cas d'erreur.
    """
    try:
        reponse = requests.get("https://api.alternative.me/fng/", timeout=10)
        reponse.raise_for_status()
        data = reponse.json()["data"][0]
        valeur = int(data["value"])
        classification = data["value_classification"]
        log(f"😱 Fear & Greed Index : {valeur}/100 ({classification})")
        return {"valeur": valeur, "classification": classification}
    except Exception as e:
        log(f"⚠️  Fear & Greed Index indisponible ({e})")
        return {}


# ─────────────────────────────────────────────────────────────
#  COMPTEUR DE TOKENS / KILL-SWITCH BUDGÉTAIRE
# ─────────────────────────────────────────────────────────────

def charger_usage_quotidien() -> dict:
    """Charge le compteur de dépenses API du jour (UTC). Remet à zéro si la date a changé."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if os.path.exists(FICHIER_TOKEN_USAGE):
        with open(FICHIER_TOKEN_USAGE) as f:
            data = json.load(f)
        if data.get("date") == today:
            return data
    return {"date": today, "cout_total_usd": 0.0, "nb_appels": 0}


def enregistrer_et_verifier_tokens(modele: str, input_tokens: int, output_tokens: int) -> float:
    """
    Comptabilise les tokens du dernier appel Claude, sauvegarde, et déclenche le
    kill-switch si la limite quotidienne est atteinte.
    """
    tarifs    = COUTS_PAR_TOKEN.get(modele, {"input": 3.00e-6, "output": 15.00e-6})
    cout_appel = input_tokens * tarifs["input"] + output_tokens * tarifs["output"]

    usage = charger_usage_quotidien()
    usage["cout_total_usd"] += cout_appel
    usage["nb_appels"]      += 1

    with open(FICHIER_TOKEN_USAGE, "w") as f:
        json.dump(usage, f, indent=2)

    pct = usage["cout_total_usd"] / LIMITE_COUT_QUOTIDIEN_USD * 100
    log(f"💳 Appel : ${cout_appel:.4f} | Jour : ${usage['cout_total_usd']:.4f} "
        f"/ ${LIMITE_COUT_QUOTIDIEN_USD:.2f} ({pct:.1f}%)")

    if usage["cout_total_usd"] >= LIMITE_COUT_QUOTIDIEN_USD:
        raise LimiteQuotidienneAtteinte(
            f"Limite quotidienne de ${LIMITE_COUT_QUOTIDIEN_USD:.2f} atteinte "
            f"(total: ${usage['cout_total_usd']:.4f} en {usage['nb_appels']} appels). "
            "Agent arrêté pour éviter une surfacturation."
        )

    return cout_appel


# ─────────────────────────────────────────────────────────────
#  ÉTAPE 3 : PRÉ-FILTRAGE DES NEWS PAR HAIKU
# ─────────────────────────────────────────────────────────────

def prefiltrer_news_haiku(articles: list, client: anthropic.Anthropic) -> tuple:
    """
    Haiku note chaque news de 0 à 3 :
      0 = pub / spam / hors sujet
      1 = news crypto générale (altcoins, NFT, DeFi sans lien BTC)
      2 = news crypto directement pertinente pour le prix du BTC
      3 = news macro importante (Fed, inflation, régulation, ETF Bitcoin)

    Retourne (articles_qualite_2_3, nb_qualite, haiku_echec).
    En cas d'erreur → ([], 0, True) : fail-safe, cycle forcé en ATTENDRE.
    """
    if not articles:
        return [], 0, False

    liste_titres = ""
    for i, article in enumerate(articles, 1):
        liste_titres += f"\n{i}. {article.get('title', 'Sans titre')}"

    prompt_filtre = f"""Note chaque news de 0 à 3 selon son importance pour le trading BTC :
  0 = pub / spam / hors sujet crypto
  1 = news crypto générale (altcoins, NFT, DeFi sans lien BTC)
  2 = news crypto directement pertinente pour le prix du BTC
  3 = news macro importante (Fed, inflation, régulation, ETF Bitcoin, crise financière)

NEWS :{liste_titres}

Réponds UNIQUEMENT avec un objet JSON valide (rien d'autre autour) :
{{
  "scores": [2, 0, 3, 1, ...]
}}
L'ordre doit correspondre exactement à la liste. Chaque score est 0, 1, 2 ou 3."""

    try:
        reponse = client.messages.create(
            model=MODELE_HAIKU,
            max_tokens=256,
            messages=[{"role": "user", "content": prompt_filtre}]
        )
        enregistrer_et_verifier_tokens(
            MODELE_HAIKU, reponse.usage.input_tokens, reponse.usage.output_tokens
        )
        contenu = reponse.content[0].text.strip()
        debut = contenu.find("{")
        fin   = contenu.rfind("}") + 1
        contenu_json = contenu[debut:fin] if debut >= 0 and fin > debut else contenu
        scores = json.loads(contenu_json).get("scores", [])

        news_qualite = [a for a, s in zip(articles, scores) if s >= 2]
        nb_qualite   = len(news_qualite)
        nb_niveau3   = sum(1 for s in scores if s == 3)
        log(f"🔍 Scoring Haiku : {nb_qualite}/{len(articles)} news qualité 2-3 "
            f"({nb_niveau3} niveau 3 macro)")
        return news_qualite, nb_qualite, False

    except LimiteQuotidienneAtteinte:
        raise
    except Exception as e:
        log(f"⚠️  Erreur scoring Haiku ({e}) — cycle forcé en ATTENDRE (fail-safe)")
        return [], 0, True


# ─────────────────────────────────────────────────────────────
#  ÉTAPE 4 : DÉCISION FINALE (Haiku ou Sonnet selon le filtre)
# ─────────────────────────────────────────────────────────────

def analyser_avec_claude(prix_btc: float, news: list, portefeuille: dict = None) -> dict:
    """
    Pipeline d'analyse en deux étapes :
    1. Haiku pré-filtre chaque news (score 0-3 : impact BTC)
    2. Si ≥ SEUIL_NEWS_PERTINENTES news pertinentes → Sonnet décide
       Sinon → Haiku décide directement

    Retourne un dictionnaire avec :
      - "score"          : de -10 (très baissier) à +10 (très haussier)
      - "recommandation" : "ACHETER", "VENDRE" ou "ATTENDRE"
      - "explication"    : 2-3 phrases d'explication
      - "modele_utilise" : nom du modèle Claude ayant pris la décision finale
    """
    usage = charger_usage_quotidien()
    if usage["cout_total_usd"] >= LIMITE_COUT_QUOTIDIEN_USD:
        raise LimiteQuotidienneAtteinte(
            f"Limite quotidienne de ${LIMITE_COUT_QUOTIDIEN_USD:.2f} déjà atteinte "
            f"(${usage['cout_total_usd']:.4f}). Aucun appel Claude émis."
        )

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # ── Étape 1 : Scoring Haiku (0-3) ──
    news_pertinentes, nb_qualite, haiku_echec = prefiltrer_news_haiku(news, client)

    # Fail-safe : si Haiku a échoué → ATTENDRE sans envoyer les news brutes
    if haiku_echec:
        log("🔒 Fail-safe Haiku actif → ATTENDRE forcé ce cycle")
        return {
            "score":           0,
            "recommandation":  "ATTENDRE",
            "explication":     "Pré-filtrage Haiku en échec — ATTENDRE par sécurité.",
            "modele_utilise":  MODELE_HAIKU,
            "nb_qualite":      0,
        }

    nb_pertinentes = nb_qualite

    # ── Choix du modèle décisionnel ──
    if nb_pertinentes >= SEUIL_NEWS_PERTINENTES:
        modele_decision = MODELE_SONNET
        log(f"📡 {nb_pertinentes} news qualité 2-3 → décision confiée à {MODELE_SONNET}")
    else:
        modele_decision = MODELE_HAIKU
        log(f"📡 {nb_pertinentes} news qualité 2-3 → décision par {MODELE_HAIKU}")

    news_pour_analyse = news_pertinentes if news_pertinentes else news

    resume_news = ""
    for i, article in enumerate(news_pour_analyse[:5], 1):
        titre    = article.get("title", "Sans titre")
        source   = article.get("source_id", "Source inconnue")
        date_pub = article.get("pubDate", "")
        resume_news += f"\n{i}. [{date_pub}] {titre} (Source: {source})"

    # ── Récupérer le contexte technique de marché ──
    marche = recuperer_donnees_marche_btc()
    var_1h  = marche.get("var_1h",  0.0)
    var_24h = marche.get("var_24h", 0.0)
    var_7j  = marche.get("var_7j",  0.0)
    min_24h = marche.get("min_24h", 0.0)
    max_24h = marche.get("max_24h", 0.0)

    # Fear & Greed Index
    fg = recuperer_fear_greed()
    fg_valeur = fg.get("valeur", None)
    fg_classification = fg.get("classification", "N/A")
    fg_str = f"{fg_valeur}/100 ({fg_classification})" if fg_valeur is not None else "N/A"

    # P&L latent si en position
    pnl_latent_str = "N/A"
    if portefeuille and portefeuille.get("en_position") and portefeuille.get("prix_achat_btc", 0) > 0:
        pnl_pct = (prix_btc / portefeuille["prix_achat_btc"] - 1) * 100
        signe = "+" if pnl_pct >= 0 else ""
        pnl_latent_str = f"{signe}{pnl_pct:.1f}%"

    contexte_technique = (
        f"Prix actuel: {prix_btc:,.0f}$ | Var 1h: {var_1h:+.2f}% | "
        f"Var 24h: {var_24h:+.2f}% | Var 7j: {var_7j:+.2f}%\n"
        f"Min/Max 24h: {min_24h:,.0f}$/{max_24h:,.0f}$ | P&L latent: {pnl_latent_str}\n"
        f"Fear & Greed Index: {fg_str}"
    )

    # ── Format de réponse JSON selon le modèle ──
    if modele_decision == MODELE_SONNET:
        format_json = """{
  "analyse_preliminaire": "<2-3 phrases de raisonnement sur les news et le contexte de marché>",
  "score": <nombre entier entre -10 et 10>,
  "recommandation": "<ACHETER ou VENDRE ou ATTENDRE>",
  "confiance": <entier de 1 (faible) à 5 (très élevé)>
}"""
    else:
        format_json = """{
  "score": <nombre entier entre -10 et 10>,
  "recommandation": "<ACHETER ou VENDRE ou ATTENDRE>",
  "explication": "<2-3 phrases expliquant ta décision en français>"
}"""

    prompt_analyse = f"""Tu es un analyste de trading crypto expérimenté et prudent.
Analyse les informations suivantes et donne une recommandation de trading BTC.

CONTEXTE DE MARCHÉ :
{contexte_technique}

NEWS QUALITÉ 2-3 ({nb_qualite} sélectionnées sur {len(news)}) :
{resume_news}

Réponds UNIQUEMENT avec un objet JSON valide dans ce format exact (rien d'autre autour) :
{format_json}

Guide de scoring :
- +9 à +10 : signal haussier exceptionnel  → ACHETER
- +5 à +8  : signal haussier notable        → ACHETER
- -4 à +4  : signal neutre                  → ATTENDRE
- -5 à -8  : signal baissier notable        → VENDRE
- -9 à -10 : signal baissier exceptionnel  → VENDRE"""

    log(f"🤖 Décision en cours avec {modele_decision}...")

    try:
        reponse = client.messages.create(
            model=modele_decision,
            max_tokens=512 if modele_decision == MODELE_HAIKU else 1024,
            messages=[{"role": "user", "content": prompt_analyse}]
        )
        enregistrer_et_verifier_tokens(
            modele_decision, reponse.usage.input_tokens, reponse.usage.output_tokens
        )
        contenu = reponse.content[0].text.strip()
        debut   = contenu.find("{")
        fin     = contenu.rfind("}") + 1
        contenu_json = contenu[debut:fin] if debut >= 0 and fin > debut else contenu

        resultat = json.loads(contenu_json)
        resultat["modele_utilise"]      = modele_decision
        resultat["nb_qualite"]          = nb_qualite
        resultat["fear_greed_valeur"]   = fg_valeur
        resultat["fear_greed_label"]    = fg_classification
        score = int(resultat.get("score", 0))
        nom_modele = "Sonnet" if modele_decision == MODELE_SONNET else "Haiku"
        log(f"📊 Score {nom_modele} : {score:+d}/10 → {resultat.get('recommandation')} "
            f"(confiance: {resultat.get('confiance', '—')}/5)")
        if resultat.get("analyse_preliminaire"):
            log(f"🧠 Raisonnement : {resultat['analyse_preliminaire']}")

    except LimiteQuotidienneAtteinte:
        raise
    except Exception as e:
        log(f"⚠️  Erreur avec {modele_decision} : {e}")
        return {
            "score":                0,
            "recommandation":       "ATTENDRE",
            "explication":          "Analyse impossible suite à une erreur technique. Par prudence : ATTENDRE.",
            "modele_utilise":       modele_decision,
            "nb_qualite":           nb_qualite,
            "fear_greed_valeur":    fg_valeur,
            "fear_greed_label":     fg_classification,
            "analyse_preliminaire": "",
            "confiance":            0,
        }

    log(f"💬 Explication : {resultat.get('explication', resultat.get('analyse_preliminaire', ''))}")
    return resultat


# ─────────────────────────────────────────────────────────────
#  ÉTAPE 6 : SIMULATION DE TRADING (paper trading — argent fictif)
# ─────────────────────────────────────────────────────────────

def simuler_transaction(portefeuille: dict, recommandation: str, prix_btc: float,
                        score_sentiment: int = 0) -> tuple:
    """
    Simule un achat ou une vente de BTC selon la recommandation de Claude.
    Applique : frais 0.1%, stop-loss -7%, take-profit +12%,
               cooldown 4h post-trade, limite 4 trades/jour,
               P&L attendu minimum 1,50 USDT.
    ⚠️  AUCUN ordre réel n'est envoyé à Binance — 100% fictif.
    """
    action       = "AUCUNE ACTION"
    raison_vente = ""
    maintenant_ts = time.time()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # ── Réinitialiser compteur quotidien si nouveau jour ──
    if portefeuille.get("date_trades", "") != today:
        portefeuille["trades_aujourd_hui"] = 0
        portefeuille["date_trades"]        = today

    # ── Vérification Stop-Loss / Take-Profit (prioritaire) ──
    if portefeuille["en_position"] and portefeuille["prix_achat_btc"] > 0:
        variation = prix_btc / portefeuille["prix_achat_btc"] - 1
        if variation <= -STOP_LOSS_PCT:
            log(f"🛑 STOP-LOSS déclenché ({variation*100:+.1f}%) → vente forcée")
            recommandation = "VENDRE"
            raison_vente   = "stop-loss"
        elif variation >= TAKE_PROFIT_PCT:
            log(f"🎯 TAKE-PROFIT déclenché ({variation*100:+.1f}%) → vente forcée")
            recommandation = "VENDRE"
            raison_vente   = "take-profit"

    # ── Vérification cooldown post-trade (4h entre trades opposés) ──
    if recommandation in ("ACHETER", "VENDRE"):
        dernier_ts   = portefeuille.get("dernier_trade_timestamp", 0.0)
        dernier_type = portefeuille.get("dernier_trade_type", "")
        temps_ecoule = maintenant_ts - dernier_ts
        # Cooldown actif seulement si c'est un trade opposé au dernier
        if (dernier_type != "" and dernier_type != recommandation
                and temps_ecoule < COOLDOWN_POST_TRADE_S
                and not raison_vente):  # SL/TP ignorent le cooldown
            restant = int((COOLDOWN_POST_TRADE_S - temps_ecoule) / 60)
            log(f"⏳ COOLDOWN actif ({dernier_type} → {recommandation}) "
                f"— encore ~{restant} min à attendre")
            recommandation = "ATTENDRE"

    # ── Vérification limite quotidienne de trades ──
    if recommandation in ("ACHETER", "VENDRE") and not raison_vente:
        nb_trades = portefeuille.get("trades_aujourd_hui", 0)
        if nb_trades >= MAX_TRADES_PAR_JOUR:
            log(f"🚫 Limite de {MAX_TRADES_PAR_JOUR} trades/jour atteinte "
                f"({nb_trades} trades) → ATTENDRE")
            recommandation = "ATTENDRE"

    # ── Vérification P&L attendu minimum (1,50 USDT = 5× frais) ──
    if recommandation == "ACHETER" and not portefeuille["en_position"]:
        capital_dispo = portefeuille["usdt_disponible"]
        pnl_attendu_achat = capital_dispo * TAKE_PROFIT_PCT - 2 * capital_dispo * FRAIS_TRADING
        if pnl_attendu_achat < MIN_PNL_ATTENDU_USDT:
            log(f"💡 P&L attendu au TP ({pnl_attendu_achat:.2f} USDT) < {MIN_PNL_ATTENDU_USDT} USDT → ATTENDRE")
            recommandation = "ATTENDRE"

    if recommandation == "VENDRE" and portefeuille["en_position"] and not raison_vente:
        btc_held = portefeuille["btc_en_stock"]
        val_brute_check = btc_held * prix_btc
        frais_v_check   = val_brute_check * FRAIS_TRADING
        pnl_si_vente    = val_brute_check - frais_v_check - btc_held * portefeuille["prix_achat_btc"]
        if pnl_si_vente < MIN_PNL_ATTENDU_USDT:
            log(f"💡 P&L si vente maintenant ({pnl_si_vente:.2f} USDT) < {MIN_PNL_ATTENDU_USDT} USDT → ATTENDRE")
            recommandation = "ATTENDRE"

    if recommandation == "ACHETER" and not portefeuille["en_position"]:
        capital    = portefeuille["usdt_disponible"]
        frais      = round(capital * FRAIS_TRADING, 4)
        btc_achete = (capital - frais) / prix_btc

        portefeuille["btc_en_stock"]   = btc_achete
        portefeuille["en_position"]    = True
        portefeuille["prix_achat_btc"] = prix_btc
        portefeuille["usdt_disponible"] = 0.0

        # Cooldown & compteur
        portefeuille["dernier_trade_timestamp"] = maintenant_ts
        portefeuille["dernier_trade_type"]      = "ACHETER"
        portefeuille["trades_aujourd_hui"]      = portefeuille.get("trades_aujourd_hui", 0) + 1
        portefeuille["date_trades"]             = today

        # Historique
        historique = charger_historique_trades()
        historique.append({
            "timestamp_ouverture": datetime.now(timezone.utc).isoformat(),
            "timestamp_fermeture": None,
            "prix_entree":         prix_btc,
            "prix_sortie":         None,
            "pnl_usdt":            None,
            "pnl_pct":             None,
            "duree_minutes":       None,
            "score_sentiment":     score_sentiment,
            "raison_sortie":       None,
            "status":              "open",
        })
        sauvegarder_historique_trades(historique)

        action = (f"📈 ACHAT FICTIF : {btc_achete:.6f} BTC "
                  f"à {prix_btc:,.2f} USDT (investi: {capital:.2f} USDT, frais: {frais:.2f} USDT)")
        log(action)

    elif recommandation == "VENDRE" and portefeuille["en_position"]:
        valeur_brute = portefeuille["btc_en_stock"] * prix_btc
        frais        = round(valeur_brute * FRAIS_TRADING, 4)
        valeur_vente = valeur_brute - frais
        cout_achat   = portefeuille["btc_en_stock"] * portefeuille["prix_achat_btc"]
        pnl_trade    = valeur_vente - cout_achat
        pnl_pct      = (valeur_vente / cout_achat - 1) * 100 if cout_achat > 0 else 0.0

        portefeuille["usdt_disponible"] = valeur_vente
        portefeuille["btc_en_stock"]    = 0.0
        portefeuille["en_position"]     = False
        portefeuille["prix_achat_btc"]  = 0.0

        # Cooldown & compteur
        portefeuille["dernier_trade_timestamp"] = maintenant_ts
        portefeuille["dernier_trade_type"]      = "VENDRE"
        portefeuille["trades_aujourd_hui"]      = portefeuille.get("trades_aujourd_hui", 0) + 1
        portefeuille["date_trades"]             = today

        # Historique : fermer le trade ouvert
        historique = charger_historique_trades()
        ts_fermeture = datetime.now(timezone.utc)
        for trade in reversed(historique):
            if trade.get("status") == "open":
                ts_ouv = datetime.fromisoformat(trade["timestamp_ouverture"])
                duree  = (ts_fermeture - ts_ouv).total_seconds() / 60
                trade["timestamp_fermeture"] = ts_fermeture.isoformat()
                trade["prix_sortie"]         = prix_btc
                trade["pnl_usdt"]            = round(pnl_trade, 4)
                trade["pnl_pct"]             = round(pnl_pct, 2)
                trade["duree_minutes"]       = round(duree, 1)
                trade["raison_sortie"]       = raison_vente if raison_vente else "signal"
                trade["status"]              = "closed"
                break
        sauvegarder_historique_trades(historique)

        prefixe = f"[{raison_vente.upper()}] " if raison_vente else ""
        signe = "+" if pnl_trade >= 0 else ""
        action = (f"{prefixe}📉 VENTE FICTIVE : {valeur_vente:.2f} USDT récupérés "
                  f"(frais: {frais:.2f} USDT | P&L trade: {signe}{pnl_trade:.2f} USDT "
                  f"/ {signe}{pnl_pct:.1f}%)")
        log(action)

    else:
        valeur_actuelle = calculer_valeur_portfolio(portefeuille, prix_btc)
        if portefeuille["en_position"]:
            variation = (prix_btc / portefeuille["prix_achat_btc"] - 1) * 100
            action = (f"⏳ EN POSITION — Valeur: {valeur_actuelle:.2f} USDT "
                      f"({variation:+.1f}% depuis achat à {portefeuille['prix_achat_btc']:,.0f} USDT)")
        else:
            action = f"⏳ EN ATTENTE — Capital disponible : {valeur_actuelle:.2f} USDT"
        log(action)

    return action, portefeuille


def calculer_valeur_portfolio(portefeuille: dict, prix_btc: float) -> float:
    """Calcule la valeur totale du portfolio fictif en USDT au prix actuel."""
    valeur = portefeuille["usdt_disponible"]
    if portefeuille["en_position"] and portefeuille["btc_en_stock"] > 0:
        valeur += portefeuille["btc_en_stock"] * prix_btc
    return valeur


def _calculer_valeur_bh(portefeuille: dict, prix_btc: float) -> float:
    """
    Valeur du portfolio Buy & Hold de référence.
    Soustrait les frais d'achat initial pour être comparable à l'agent.
    """
    prix_initial = portefeuille.get("prix_initial_bh", 0.0)
    if prix_initial <= 0 or prix_btc <= 0:
        return CAPITAL_DEPART_USDT
    return (CAPITAL_DEPART_USDT * (1 - FRAIS_TRADING)) / prix_initial * prix_btc


# ─────────────────────────────────────────────────────────────
#  ÉTAPE 5 : ENREGISTREMENT DANS EXCEL
# ─────────────────────────────────────────────────────────────

EN_TETES_EXCEL = [
    "Date",
    "Heure",
    "Prix BTC (USDT)",
    "Nb news analysées",
    "Titres des news",
    "Score sentiment",
    "Modèle Claude utilisé",
    "Recommandation",
    "Action simulée",
    "Capital fictif (USDT)",
    "P&L total (USDT)",
    "Valeur B&H (USDT)",
    "P&L B&H (USDT)",
    "News qualité 2-3",
    "Raisonnement Claude",
    "Fear & Greed",
]


def initialiser_excel():
    """Crée le fichier Excel avec les en-têtes si il n'existe pas encore."""
    bold = openpyxl.styles.Font(bold=True)

    if not os.path.exists(FICHIER_EXCEL):
        classeur = openpyxl.Workbook()
        feuille  = classeur.active
        feuille.title = "Journal de Trading"
        feuille.append(EN_TETES_EXCEL)
        for cellule in feuille[1]:
            cellule.font = bold
        classeur.save(FICHIER_EXCEL)
        log(f"📁 Fichier Excel créé : {FICHIER_EXCEL}")
        return

    classeur = openpyxl.load_workbook(FICHIER_EXCEL)
    feuille  = classeur.active
    en_tetes_actuels = [c.value for c in feuille[1]]
    if "Valeur B&H (USDT)" not in en_tetes_actuels:
        col = len(en_tetes_actuels) + 1
        cell_bh  = feuille.cell(row=1, column=col,   value="Valeur B&H (USDT)")
        cell_pnl = feuille.cell(row=1, column=col+1, value="P&L B&H (USDT)")
        cell_bh.font  = bold
        cell_pnl.font = bold
        classeur.save(FICHIER_EXCEL)
        en_tetes_actuels = [c.value for c in feuille[1]]
        log("📁 Colonnes Buy & Hold ajoutées au fichier Excel existant.")
    if "News qualité 2-3" not in en_tetes_actuels:
        col = len(en_tetes_actuels) + 1
        cell = feuille.cell(row=1, column=col, value="News qualité 2-3")
        cell.font = bold
        classeur.save(FICHIER_EXCEL)
        en_tetes_actuels = [c.value for c in feuille[1]]
        log("📁 Colonne 'News qualité 2-3' ajoutée au fichier Excel existant.")
    for col_name in ("Raisonnement Claude", "Fear & Greed"):
        if col_name not in en_tetes_actuels:
            col = len(en_tetes_actuels) + 1
            cell = feuille.cell(row=1, column=col, value=col_name)
            cell.font = bold
            classeur.save(FICHIER_EXCEL)
            en_tetes_actuels = [c.value for c in feuille[1]]
            log(f"📁 Colonne '{col_name}' ajoutée au fichier Excel existant.")


def enregistrer_dans_excel(prix_btc: float, news: list, analyse: dict,
                            action: str, valeur_portfolio: float, valeur_bh: float):
    """Ajoute une nouvelle ligne dans le fichier Excel avec toutes les données du cycle."""
    pnl_total = valeur_portfolio - CAPITAL_DEPART_USDT
    pnl_bh    = valeur_bh - CAPITAL_DEPART_USDT

    titres_news = " | ".join(
        article.get("title", "Sans titre")[:80]
        for article in news[:5]
    ) or "Aucune news"

    fg_valeur = analyse.get("fear_greed_valeur", None)
    fg_label  = analyse.get("fear_greed_label", "")
    fg_excel  = f"{fg_valeur}/100 ({fg_label})" if fg_valeur is not None else "N/A"

    maintenant = datetime.now(timezone.utc)
    nouvelle_ligne = [
        maintenant.strftime("%Y-%m-%d"),
        maintenant.strftime("%H:%M:%S"),
        round(prix_btc, 2),
        len(news),
        titres_news,
        analyse.get("score", 0),
        analyse.get("modele_utilise", "N/A"),
        analyse.get("recommandation", "ATTENDRE"),
        action,
        round(valeur_portfolio, 2),
        round(pnl_total, 2),
        round(valeur_bh, 2),
        round(pnl_bh, 2),
        analyse.get("nb_qualite", 0),
        analyse.get("analyse_preliminaire", ""),
        fg_excel,
    ]

    classeur = openpyxl.load_workbook(FICHIER_EXCEL)
    feuille  = classeur.active
    feuille.append(nouvelle_ligne)
    classeur.save(FICHIER_EXCEL)

    signe_agent = "+" if pnl_total >= 0 else ""
    signe_bh    = "+" if pnl_bh >= 0 else ""
    log(f"💾 Excel — P&L agent : {signe_agent}{pnl_total:.2f} USDT | P&L B&H : {signe_bh}{pnl_bh:.2f} USDT")


# ═════════════════════════════════════════════════════════════
#  CYCLE PRINCIPAL
# ═════════════════════════════════════════════════════════════

def executer_un_cycle(portefeuille: dict) -> dict:
    """
    Exécute UN cycle complet de l'agent :
    1. Vérifier les nouvelles news (filtre âge 6h)
    2. Si 0 news récentes → ATTENDRE sans appeler Claude
    3. Récupérer le prix BTC
    4. Analyser avec Claude (Haiku ou Sonnet selon le signal)
    5. Simuler une transaction (paper trading)
    6. Enregistrer dans Excel
    7. Retourner le portefeuille mis à jour
    """
    separateur = "─" * 55
    log(separateur)
    log("🔄 Nouveau cycle d'analyse")
    log(separateur)

    # ── Étape 1 : Vérifier les nouvelles news ──
    nouvelles_news = recuperer_nouvelles_news()

    if not nouvelles_news:
        log("💤 Pas de nouvelles news ce cycle — écriture heartbeat dans Excel.")
        prix_btc = recuperer_prix_btc()
        valeur_portfolio = calculer_valeur_portfolio(portefeuille, prix_btc)
        valeur_bh = _calculer_valeur_bh(portefeuille, prix_btc)
        enregistrer_dans_excel(
            prix_btc, [],
            {"score": 0, "recommandation": "ATTENDRE", "modele_utilise": "—", "nb_qualite": 0},
            "ATTENDRE",
            valeur_portfolio,
            valeur_bh,
        )
        return portefeuille

    # ── Étape 2 : Récupérer le prix BTC ──
    prix_btc = recuperer_prix_btc()
    if prix_btc == 0.0:
        log("⚠️  Prix BTC indisponible — cycle annulé par sécurité.")
        return portefeuille

    # ── Initialiser le prix de référence Buy & Hold au premier cycle valide ──
    if portefeuille.get("prix_initial_bh", 0.0) == 0.0:
        portefeuille["prix_initial_bh"] = prix_btc
        log(f"📊 Prix de référence Buy & Hold initialisé : {prix_btc:,.2f} USDT")

    # ── Étapes 3 & 4 : Analyser avec Claude (avec contexte portefeuille) ──
    analyse = analyser_avec_claude(prix_btc, nouvelles_news, portefeuille=portefeuille)

    # ── Étape 6 : Simuler une transaction ──
    recommandation = analyse.get("recommandation", "ATTENDRE")

    # Hausser le seuil de vente : score ≤ -5 ET Sonnet requis (pas Haiku seul)
    if recommandation == "VENDRE":
        score_analyse  = analyse.get("score", 0)
        modele_analyse = analyse.get("modele_utilise", "")
        if score_analyse > -5 or modele_analyse != MODELE_SONNET:
            log(f"⛔ Vente bloquée : score {score_analyse:+d} (requis ≤ -5) "
                f"ou modèle non Sonnet ({modele_analyse}) → ATTENDRE")
            recommandation = "ATTENDRE"

    action, portefeuille = simuler_transaction(
        portefeuille, recommandation, prix_btc,
        score_sentiment=analyse.get("score", 0)
    )

    # ── Sauvegarder le portefeuille ──
    sauvegarder_portfolio(portefeuille)

    # ── Étape 5 : Enregistrer dans Excel ──
    valeur_portfolio = calculer_valeur_portfolio(portefeuille, prix_btc)
    valeur_bh = _calculer_valeur_bh(portefeuille, prix_btc)
    enregistrer_dans_excel(prix_btc, nouvelles_news, analyse, action, valeur_portfolio, valeur_bh)

    pnl    = valeur_portfolio - CAPITAL_DEPART_USDT
    pnl_bh = valeur_bh - CAPITAL_DEPART_USDT
    signe    = "+" if pnl >= 0 else ""
    signe_bh = "+" if pnl_bh >= 0 else ""
    log(f"💼 Agent : {valeur_portfolio:.2f} USDT (P&L: {signe}{pnl:.2f}) | "
        f"B&H : {valeur_bh:.2f} USDT (P&L: {signe_bh}{pnl_bh:.2f})")

    return portefeuille


# ═════════════════════════════════════════════════════════════
#  POINT D'ENTRÉE PRINCIPAL
# ═════════════════════════════════════════════════════════════

def main():
    """
    Lance l'agent de trading.

    Deux modes de fonctionnement :
    - Mode local (Mac)           : boucle infinie, attend 15 min entre chaque cycle
    - Mode GitHub Actions (--once) : exécute UN seul cycle puis s'arrête
    """
    mode_unique = "--once" in sys.argv

    log("=" * 55)
    log("🚀 DÉMARRAGE DE L'AGENT DE TRADING CRYPTO")
    if mode_unique:
        log("   Mode : GitHub Actions (un seul cycle)")
    else:
        log(f"   Mode : local (boucle toutes les {INTERVALLE_SECONDES // 60} min)")
    log(f"   Capital fictif de départ : {CAPITAL_DEPART_USDT:.2f} USDT")
    log(f"   Fichier Excel            : {FICHIER_EXCEL}")
    log(f"   Seuil news qualité 2-3   : ≥{SEUIL_NEWS_PERTINENTES} → Sonnet décide")
    log(f"   Frais trading            : {FRAIS_TRADING*100:.1f}% par trade")
    log(f"   Stop-loss                : -{STOP_LOSS_PCT*100:.0f}%")
    log(f"   Take-profit              : +{TAKE_PROFIT_PCT*100:.0f}%")
    log(f"   Cooldown post-trade      : {COOLDOWN_POST_TRADE_S // 3600}h")
    log(f"   Max trades/jour          : {MAX_TRADES_PAR_JOUR}")
    log(f"   Filtre âge news          : {FILTRE_AGE_MAX_SECONDES // 3600}h max")
    log(f"   Limite coût API/jour     : ${LIMITE_COUT_QUOTIDIEN_USD:.2f}")
    usage_actuel = charger_usage_quotidien()
    log(f"   Dépense API aujourd'hui  : ${usage_actuel['cout_total_usd']:.4f} "
        f"({usage_actuel['nb_appels']} appels)")
    log("=" * 55)
    log("")

    if not ANTHROPIC_API_KEY:
        log("❌ ERREUR : Clé API manquante : ANTHROPIC_API_KEY")
        log("   En local : remplis le fichier .env")
        log("   Sur GitHub : ajoute le secret dans Settings → Secrets → Actions")
        sys.exit(1)

    if not BINANCE_API_KEY:
        log("ℹ️  Pas de clé Binance — le prix BTC sera quand même récupéré (lecture publique).")

    initialiser_excel()
    portefeuille = charger_portfolio()

    if portefeuille["en_position"]:
        log(f"📂 Portfolio chargé — En position BTC ({portefeuille['btc_en_stock']:.6f} BTC)")
    else:
        log(f"📂 Portfolio chargé — USDT disponible : {portefeuille['usdt_disponible']:.2f}")

    nb_trades = portefeuille.get("trades_aujourd_hui", 0)
    if nb_trades > 0:
        log(f"📊 Trades aujourd'hui : {nb_trades}/{MAX_TRADES_PAR_JOUR}")

    # ════════════════════════════════════════════
    #  MODE GITHUB ACTIONS : un seul cycle puis exit
    # ════════════════════════════════════════════
    if mode_unique:
        log("\n▶  Exécution du cycle unique (mode GitHub Actions)...")
        try:
            executer_un_cycle(portefeuille)
        except LimiteQuotidienneAtteinte as e:
            log(f"🚨 KILL-SWITCH BUDGÉTAIRE : {e}")
            sys.exit(1)
        except Exception as e:
            log(f"❌ Erreur : {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
        log("✅ Cycle terminé. GitHub Actions sauvegardera les fichiers.")
        return

    # ════════════════════════════════════════════
    #  MODE LOCAL : boucle infinie toutes les 15 min
    # ════════════════════════════════════════════
    numero_cycle = 0
    while True:
        numero_cycle += 1
        log(f"\n{'═' * 55}")
        log(f"CYCLE N°{numero_cycle}")

        try:
            portefeuille = executer_un_cycle(portefeuille)
        except KeyboardInterrupt:
            log("\n⛔ Arrêt demandé (Ctrl+C). À bientôt !")
            break
        except LimiteQuotidienneAtteinte as e:
            log(f"🚨 KILL-SWITCH BUDGÉTAIRE : {e}")
            break
        except Exception as e:
            log(f"❌ Erreur inattendue : {e}")
            import traceback
            traceback.print_exc()
            log("   L'agent continue malgré l'erreur...")

        log(f"\n⏰ Prochain cycle dans {INTERVALLE_SECONDES // 60} minutes...")
        try:
            time.sleep(INTERVALLE_SECONDES)
        except KeyboardInterrupt:
            log("\n⛔ Arrêt demandé. À bientôt !")
            break


if __name__ == "__main__":
    main()
